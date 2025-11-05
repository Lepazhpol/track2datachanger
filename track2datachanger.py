import re
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import uuid
import csv
from openpyxl import load_workbook


# -----------------------
# Утилиты
# -----------------------

def luhn_check_digit(pan_without_last: str) -> int:
    total = 0
    alt = True
    for d in reversed(pan_without_last):
        n = int(d)
        if alt:
            n *= 2
            if n > 9:
                n -= 9
        total += n
        alt = not alt
    return (10 - (total % 10)) % 10


def luhn_check(pan: str) -> bool:
    total = 0
    alt = False
    for d in reversed(pan):
        n = int(d)
        if alt:
            n *= 2
            if n > 9:
                n -= 9
        total += n
        alt = not alt
    return total % 10 == 0


def interpret_service_code(sc: str) -> str:
    """Человекочитаемая расшифровка service code."""
    if not sc or len(sc) != 3 or not sc.isdigit():
        return "Invalid service code"
    a, b, c = sc[0], sc[1], sc[2]
    first_map = {
        '1': "International interchange OK",
        '2': "IC (chip) preferred",
        '5': "National use only",
        '6': "National use only, IC preferred",
        '7': "No interchange (private network)",
        '9': "Test card",
    }
    second_map = {
        '0': "Normal processing",
        '2': "Authorization via online required",
        '4': "Online except bilateral agreement",
    }
    third_map = {
        '0': "No restrictions (PIN OK)",
        '1': "No restrictions (PIN not required)",
        '2': "Goods/services only (no cash)",
        '3': "ATM only (PIN required)",
        '4': "Cash only",
        '5': "Goods/services only (PIN required)",
        '6': "No restrictions (use PIN where feasible)",
        '7': "Goods/services only (use PIN feasible)",
    }
    f = first_map.get(a, "Unknown")
    s = second_map.get(b, "Unknown")
    t = third_map.get(c, "Unknown")
    return f"{sc}: {f}; {s}; {t}"


# -----------------------
# Парсер Track2
# -----------------------

def parse_track2(track: str, bin_map=None):
    track = track.strip()
    sep = '=' if '=' in track else ('D' if 'D' in track else None)
    pattern = re.compile(r'^(?P<PAN>\d{12,19})[=D](?P<EXP>\d{4})(?P<SVC>\d{3})(?P<DISCR>\d*)$')
    m = pattern.match(track)

    result = {
        'raw': track,
        'PAN': '',
        'expiry': '',
        'expired': '',
        'svc': '',
        'svc_human': '',
        'discr_len': '',
        'luhn_ok': '',
        'status': '',
        'details': '',
        'correct_pan': None,
        'track_type': 'Unknown',
        'bin': '',
        'brand': 'Unknown',
    }

    if not m:
        result['status'] = '❌ Format'
        result['details'] = 'Неверный формат Track2.'
        return result

    pan = m.group('PAN')
    exp = m.group('EXP')
    svc = m.group('SVC')
    discr = m.group('DISCR')

    result['PAN'] = pan
    result['svc'] = svc
    result['svc_human'] = interpret_service_code(svc)
    result['discr_len'] = str(len(discr))
    result['track_type'] = "EMV/Chip ('D')" if sep == 'D' else "Magstripe ('=')"
    result['bin'] = pan[:6] if len(pan) >= 6 else pan

    # Определяем бренд только по импортированному справочнику
    if bin_map and result['bin'] in bin_map:
        result['brand'] = bin_map[result['bin']]

    # Проверка Луна
    if luhn_check(pan):
        result['luhn_ok'] = 'OK'
    else:
        result['luhn_ok'] = 'FAIL'
        correct_digit = luhn_check_digit(pan[:-1])
        result['correct_pan'] = pan[:-1] + str(correct_digit)
        result['status'] = '❌ Luhn fail'
        result['details'] = f'Неверная контрольная сумма. Правильный check digit: {correct_digit}'
        return result

    # Проверка срока действия
    try:
        yy, mm = int(exp[:2]), int(exp[2:])
        year = 2000 + yy if yy < 80 else 1900 + yy
        expired = datetime(year, mm, 1) < datetime.now().replace(day=1)
        result['expiry'] = f'{exp[:2]}/{exp[2:]}'
        result['expired'] = 'YES' if expired else 'NO'
        if expired:
            result['status'] = '❗Expired'
            result['details'] = 'Срок действия истёк.'
            return result
    except Exception:
        result['status'] = '❌ Exp parse'
        result['details'] = 'Ошибка разбора срока действия.'
        return result

    result['status'] = '✅ OK'
    result['details'] = 'Все проверки пройдены успешно.'
    return result


# -----------------------
# GUI
# -----------------------

class Track2CheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Track2 Checker")
        self.bin_map = {}  # справочник BIN -> brand

        window_width = 1000
        window_height = 650
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        root.geometry(f"{window_width}x{window_height}+{(sw - window_width)//2}+{(sh - window_height)//2}")

        lbl = tk.Label(root, text="Введите или вставьте Track2Data (по одному на строку):")
        lbl.pack(pady=(10, 0))

        self.input_text = tk.Text(root, height=6, width=120, font=("Consolas", 11))
        self.input_text.pack(padx=10, pady=5)

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=(0, 6))

        tk.Button(btn_frame, text="Проверить", command=self.check_tracks,
                  bg="#0078d7", fg="white", activebackground="#005a9e",
                  activeforeground="white", font=("Segoe UI", 10, "bold"), relief="flat").pack(side='left', padx=5)
        tk.Button(btn_frame, text="Импорт .txt", command=self.import_file).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Импорт BIN .xlsx", command=self.import_bin_xlsx).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Экспорт CSV", command=self.export_csv).pack(side='left', padx=5)

        columns = ("PAN", "Expiry", "Luhn", "Expired", "Svc", "Discr", "Status", "BIN", "Brand", "TrackType")
        self.tree = ttk.Treeview(root, columns=columns, show='headings')
        headers = [
            ("PAN", "PAN", 200),
            ("Expiry", "Expiry", 90),
            ("Luhn", "Luhn", 60),
            ("Expired", "Expired?", 80),
            ("Svc", "Svc", 160),
            ("Discr", "Discr.len", 80),
            ("Status", "Status", 120),
            ("BIN", "BIN", 90),
            ("Brand", "Brand", 120),
            ("TrackType", "TrackType", 180),
        ]
        for col, text, w in headers:
            self.tree.heading(col, text=text)
            self.tree.column(col, width=w, anchor="center")
        self.tree.pack(fill='both', expand=True, padx=10, pady=(0, 4))

        tk.Label(root, text="Двойной клик для деталей", font=("Segoe UI", 9), fg="#777").pack(pady=(0, 4))

        self.summary_var = tk.StringVar(value="OK: 0   Expired: 0   Luhn fail: 0   Format error: 0")
        tk.Label(root, textvariable=self.summary_var, font=("Segoe UI", 10), fg="#ccc").pack(pady=(0, 10))

        self.tree.bind("<Double-1>", self.show_details)

    # --- Импорт BIN-справочника
    def import_bin_xlsx(self):
        path = filedialog.askopenfilename(title="Импорт BIN справочника", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            wb = load_workbook(path)
            ws = wb.active
            new_map = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2:
                    continue
                brand, bin_value = str(row[0]).strip(), str(row[1]).strip()
                if brand and bin_value:
                    new_map[bin_value] = brand
            self.bin_map = new_map
            messagebox.showinfo("Импорт BIN", f"Импортировано {len(new_map)} BIN-записей.")
        except Exception as e:
            messagebox.showerror("Ошибка импорта", str(e))

    def import_file(self):
        path = filedialog.askopenfilename(title="Импорт .txt", filetypes=[("Text files", "*.txt")])
        if not path:
            return
        with open(path, "r", encoding="utf-8") as f:
            data = f.read()
        self.input_text.delete("1.0", "end")
        self.input_text.insert("1.0", data)

    def export_csv(self):
        if not self.tree.get_children():
            messagebox.showinfo("Экспорт CSV", "Нет данных для экспорта.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if not path:
            return
        with open(path, "w", newline='', encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([self.tree.heading(c)["text"] for c in self.tree["columns"]])
            for row in self.tree.get_children():
                writer.writerow(self.tree.item(row)["values"])
        messagebox.showinfo("Экспорт CSV", f"Сохранено {len(self.tree.get_children())} строк.")

    def check_tracks(self):
        self.tree.delete(*self.tree.get_children())
        lines = [x.strip() for x in self.input_text.get("1.0", "end").splitlines() if x.strip()]
        if not lines:
            messagebox.showinfo("Track2 Checker", "Введите хотя бы один трек.")
            return

        c_ok = c_exp = c_luhn = c_fmt = 0
        for t in lines:
            res = parse_track2(t, self.bin_map)
            iid = str(uuid.uuid4())
            self.tree.insert("", "end", values=(
                res['PAN'], res['expiry'], res['luhn_ok'], res['expired'], res['svc_human'],
                res['discr_len'], res['status'], res['bin'], res['brand'], res['track_type']
            ), iid=iid, tags=(str(res),))

            s = res['status']
            if s.startswith('✅'):
                c_ok += 1
            elif s.startswith('❗'):
                c_exp += 1
            elif 'Luhn' in s:
                c_luhn += 1
            elif 'Format' in s:
                c_fmt += 1
        self.summary_var.set(f"OK: {c_ok}   Expired: {c_exp}   Luhn fail: {c_luhn}   Format error: {c_fmt}")

    def show_details(self, event):
        item_id = self.tree.focus()
        if not item_id:
            return
        vals = self.tree.item(item_id)["values"]
        text = (
            f"PAN: {vals[0]}\n"
            f"Expiry: {vals[1]}\n"
            f"Luhn: {vals[2]}\n"
            f"Expired?: {vals[3]}\n"
            f"Service code: {vals[4]}\n"
            f"Discr.len: {vals[5]}\n"
            f"Status: {vals[6]}\n"
            f"BIN: {vals[7]}\n"
            f"Brand: {vals[8]}\n"
            f"Track type: {vals[9]}"
        )
        tag = self.tree.item(item_id, "tags")[0]
        if "correct_pan" in tag:
            m = re.search(r"'correct_pan': '(\d+)'", tag)
            if m:
                text += f"\n\nCorrect PAN (valid Luhn): {m.group(1)}"
        messagebox.showinfo("Детали проверки", text)


if __name__ == "__main__":
    root = tk.Tk()
    app = Track2CheckerApp(root)
    root.mainloop()